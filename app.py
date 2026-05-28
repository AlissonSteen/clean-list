import io
import os
import re
import json
import uuid
import csv
from collections import OrderedDict
from flask import Flask, request, jsonify, send_file, render_template, Response, stream_with_context
import pandas as pd
import numpy as np

app = Flask(__name__)
UPLOAD_CACHE_MAX = 8
RESULT_CACHE_MAX = 16
UPLOAD_CACHE = OrderedDict()
RESULT_CACHE = OrderedDict()

# ─── Phone helpers ─────────────────────────────────────────────────────────────

def extract_phones(raw: str):
    raw = str(raw).strip()
    parts = re.split(r'[;|/\\\n]+|(?<!\d),(?!\d{3})', raw)
    cleaned_parts = [only_digits(p.strip()) for p in parts]
    return [digits for digits in cleaned_parts if digits]


def parse_list_cell(raw: str) -> list[str]:
    text = str(raw or '').strip()
    if not text:
        return []

    normalized = (
        text.replace('“', '"')
        .replace('”', '"')
        .replace('’', "'")
        .replace('‘', "'")
    )

    try:
        parsed = next(csv.reader([normalized], skipinitialspace=True))
        values = [re.sub(r'\s+', ' ', value).strip().strip('"').strip("'").strip() for value in parsed]
        values = [value for value in values if value]
        if values:
            return values
    except Exception:
        pass

    fallback = [re.sub(r'\s+', ' ', value).strip().strip('"').strip("'").strip()
                for value in re.split(r'[;|/\n]+', normalized)]
    return [value for value in fallback if value]


def only_digits(val: str) -> str:
    text = str(val).strip()
    if re.fullmatch(r'\d+\.0+', text):
        text = text.split('.', 1)[0]
    return re.sub(r'[^\d]', '', text)


def is_repeated_digits(digits: str) -> bool:
    return bool(digits) and len(set(digits)) == 1


def is_valid_cpf(digits: str) -> bool:
    if len(digits) != 11 or is_repeated_digits(digits):
        return False

    total = sum(int(digits[i]) * (10 - i) for i in range(9))
    check_1 = (total * 10 % 11) % 10
    total = sum(int(digits[i]) * (11 - i) for i in range(10))
    check_2 = (total * 10 % 11) % 10
    return digits[-2:] == f'{check_1}{check_2}'


def is_valid_cnpj(digits: str) -> bool:
    if len(digits) != 14 or is_repeated_digits(digits):
        return False

    def _calc(base: str, weights: list[int]) -> str:
        total = sum(int(d) * w for d, w in zip(base, weights))
        remainder = total % 11
        return '0' if remainder < 2 else str(11 - remainder)

    check_1 = _calc(digits[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    check_2 = _calc(digits[:12] + check_1, [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    return digits[-2:] == f'{check_1}{check_2}'


def looks_like_document_id(val: str) -> bool:
    digits = only_digits(val)
    return is_valid_cpf(digits) or is_valid_cnpj(digits)


def is_document_column_name(col_name: str) -> bool:
    normalized = re.sub(r'[^a-z0-9]', '', str(col_name).lower())
    return any(token in normalized for token in ('cpf', 'cnpj', 'documento', 'doc', 'rg'))


def normalize_phone(digits: str):
    digits = only_digits(digits)
    if not digits:
        return None
    while digits.startswith('55') and len(digits) > 11:
        digits = digits[2:]
    if len(digits) < 8 or len(digits) > 11:
        return None
    ddd = int(digits[:2])
    if not 11 <= ddd <= 99:
        return None
    number = digits[2:]

    if len(number) in (6, 7):
        return f'55{ddd:02d}{number}'

    if ddd < 28:
        if len(number) == 8:
            number = f'9{number}'
        elif len(number) == 9:
            if not number.startswith('9'):
                return None
        else:
            return None
    else:
        if len(number) == 9:
            if number.startswith('9'):
                number = number[1:]
            else:
                return None
        elif len(number) != 8:
            return None

    return f'55{ddd:02d}{number}'


def looks_like_phone(val: str, reject_documents: bool = True) -> str | None:
    """Return normalized phone if val looks like a phone number, else None."""
    digits = only_digits(val)
    normalized = normalize_phone(digits)
    if not normalized:
        return None
    if reject_documents and looks_like_document_id(digits):
        return None
    return normalized


# ─── Name helpers ──────────────────────────────────────────────────────────────

def split_name(full: str):
    parts = str(full).strip().split()
    if not parts:
        return '', ''
    return parts[0].capitalize(), ' '.join(p.capitalize() for p in parts[1:])


def normalize_person_name(full: str) -> str:
    text = re.sub(r'\s+', ' ', str(full or '')).strip().strip('"').strip("'").strip()
    if not text:
        return ''
    return ' '.join(part.capitalize() for part in text.split())


def infer_phone_column(df: pd.DataFrame) -> str | None:
    best_col = None
    best_score = 0.0

    for col in df.columns:
        if is_document_column_name(col):
            continue
        series = df[col].fillna('').astype(str).str.strip()
        sample = series[series.ne('')].head(200)
        if sample.empty:
            continue

        phone_hits = sample.apply(lambda val: looks_like_phone(val) is not None).sum()
        score = phone_hits / len(sample)

        if phone_hits >= 2 and score > best_score:
            best_col = col
            best_score = score

    return best_col if best_score >= 0.35 else None


def infer_name_column(df: pd.DataFrame, phone_col: str | None = None) -> str | None:
    best_col = None
    best_score = 0.0

    for col in df.columns:
        if phone_col and col == phone_col:
            continue

        series = df[col].fillna('').astype(str).str.strip()
        sample = series[series.ne('')].head(200)
        if sample.empty:
            continue

        def _looks_like_name(val: str) -> bool:
            return looks_like_person_name_value(val, allow_single_word=True)

        name_hits = sample.apply(_looks_like_name).sum()
        score = name_hits / len(sample)

        if name_hits >= 2 and score > best_score:
            best_col = col
            best_score = score

    return best_col if best_score >= 0.3 else None


def infer_columns(df: pd.DataFrame) -> dict:
    phone_col = infer_phone_column(df)
    name_col = infer_name_column(df, phone_col)
    return {
        'name_col': name_col,
        'phone_col': phone_col,
    }


def is_known_header_name(value: str) -> bool:
    normalized = re.sub(r'[^a-z0-9]', '', str(value).lower())
    known_tokens = (
        'nome', 'name', 'telefone', 'phone', 'cel', 'celular', 'whatsapp',
        'cpf', 'cnpj', 'email', 'mail', 'contato', 'cliente', 'titular',
        'valor', 'data', 'vencimento', 'parcela', 'documento', 'doc',
    )
    return any(token in normalized for token in known_tokens)


def looks_like_person_name_value(value: str, allow_single_word: bool = False) -> bool:
    text = str(value or '').strip()
    if not text or looks_like_phone(text, reject_documents=False) or looks_like_document_id(text):
        return False
    if is_known_header_name(text):
        return False

    cleaned = re.sub(r'[^A-Za-zÀ-ÿ\s]', ' ', text).strip()
    parts = [part for part in cleaned.split() if part]
    if not parts:
        return False

    letters = re.sub(r'[^A-Za-zÀ-ÿ]', '', cleaned)
    if len(parts) >= 2:
        return len(letters) >= 5

    return allow_single_word and len(letters) >= 3 and text[:1].isupper()


def looks_like_headerless_dataframe(df: pd.DataFrame) -> bool:
    if df.empty or len(df.columns) == 0:
        return False

    header_values = [str(col).strip() for col in df.columns]
    populated_header_values = [value for value in header_values if value]
    if not populated_header_values:
        return False
    if any(is_known_header_name(value) for value in populated_header_values):
        return False

    data_like_hits = 0
    for value in populated_header_values:
        if looks_like_phone(value, reject_documents=False):
            data_like_hits += 1
            continue
        if looks_like_document_id(value):
            data_like_hits += 1
            continue
        if looks_like_person_name_value(value, allow_single_word=True):
            data_like_hits += 1

    required_hits = max(1, (len(populated_header_values) + 1) // 2)
    return data_like_hits >= required_hits


def generate_headers_from_samples(values: list[str]) -> list[str]:
    generated = []
    counters = {'nome': 0, 'telefone': 0, 'cpf': 0, 'coluna': 0}

    for value in values:
        text = str(value).strip()
        if looks_like_phone(text, reject_documents=False):
            base = 'telefone'
        elif looks_like_document_id(text):
            base = 'cpf'
        elif re.search(r'[A-Za-zÀ-ÿ]', text):
            base = 'nome'
        else:
            base = 'coluna'

        counters[base] += 1
        suffix = '' if counters[base] == 1 else f'_{counters[base]}'
        generated.append(f'{base}{suffix}')

    return generated


def restore_headerless_first_row(df: pd.DataFrame) -> pd.DataFrame:
    if not looks_like_headerless_dataframe(df):
        return df

    first_row = [str(col).strip() for col in df.columns]
    restored = pd.concat(
        [pd.DataFrame([first_row], columns=df.columns), df.reset_index(drop=True)],
        ignore_index=True,
    )
    restored.columns = generate_headers_from_samples(first_row)
    return restored.astype(str)


def normalize_imported_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.astype(str)

    normalized = df.fillna('').astype(str).copy()
    normalized = normalized.loc[normalized.apply(lambda row: row.str.strip().ne('').any(), axis=1)]
    if normalized.empty:
        return normalized.reset_index(drop=True)

    non_empty_cols = normalized.apply(lambda col: col.str.strip().ne('').any())
    normalized = normalized.loc[:, non_empty_cols]
    if normalized.empty:
        return normalized.reset_index(drop=True)

    new_columns = []
    used_names: dict[str, int] = {}

    for idx, col in enumerate(normalized.columns):
        label = str(col).strip()
        if not label or label.lower().startswith('unnamed:'):
            sample = normalized.iloc[:, idx].astype(str).str.strip()
            sample_values = sample[sample.ne('')].head(1).tolist()
            if sample_values:
                label = generate_headers_from_samples(sample_values)[0]
            else:
                label = 'coluna'

        count = used_names.get(label, 0) + 1
        used_names[label] = count
        if count > 1:
            label = f'{label}_{count}'
        new_columns.append(label)

    normalized.columns = new_columns
    return normalized.reset_index(drop=True)


def build_download_name(original_name: str, fmt: str) -> str:
    base_name = os.path.basename(original_name or 'arquivo')
    root, _ = os.path.splitext(base_name)
    root = root or 'arquivo'
    ext = 'csv' if fmt == 'csv' else 'xlsx'
    return f'{root}_higienizado.{ext}'


def build_split_download_name(original_name: str, fmt: str, part_index: int, total_parts: int) -> str:
    base_name = os.path.basename(original_name or 'arquivo')
    root, _ = os.path.splitext(base_name)
    root = root or 'arquivo'
    ext = 'csv' if fmt == 'csv' else 'xlsx'
    width = max(2, len(str(total_parts)))
    return f'{root}_higienizado_parte_{part_index:0{width}d}_de_{total_parts:0{width}d}.{ext}'


def split_dataframe_evenly(df: pd.DataFrame, total_parts: int) -> list[pd.DataFrame]:
    total_rows = len(df)
    total_parts = max(1, int(total_parts))
    if total_rows > 0:
        total_parts = min(total_parts, total_rows)
    if total_parts == 1 or df.empty:
        return [df]

    base_size, remainder = divmod(total_rows, total_parts)
    slices = []
    start = 0
    for part_idx in range(total_parts):
        part_size = base_size + (1 if part_idx < remainder else 0)
        end = start + part_size
        slices.append(df.iloc[start:end].copy())
        start = end
    return slices


def make_error_payload(title: str, message: str, hint: str = '', kind: str = 'error',
                       details: str = '') -> dict:
    payload = {
        'title': title,
        'message': message,
        'kind': kind,
    }
    if hint:
        payload['hint'] = hint
    if details:
        payload['details'] = details
    return payload


def classify_exception(exc: Exception, context: str = 'general') -> dict:
    message = str(exc).strip()

    if 'Nenhum arquivo enviado' in message:
        return make_error_payload(
            'Arquivo não enviado',
            'Nenhum arquivo foi enviado para o sistema.',
            'Selecione um arquivo antes de continuar.'
        )

    if 'Formato não suportado' in message:
        return make_error_payload(
            'Formato não suportado',
            'Não conseguimos ler esse tipo de arquivo.',
            'Use um arquivo .xlsx, .xls, .csv ou .pdf.'
        )

    if 'Arquivo sem colunas legíveis' in message:
        return make_error_payload(
            'Arquivo sem colunas legíveis',
            'O arquivo foi recebido, mas não encontramos colunas válidas para montar a tabela.',
            'Revise o arquivo de origem e confira se ele tem cabeçalhos e dados estruturados.'
        )

    if isinstance(exc, pd.errors.EmptyDataError):
        return make_error_payload(
            'Arquivo vazio',
            'O arquivo não contém dados suficientes para leitura.',
            'Verifique se ele possui linhas e colunas preenchidas.'
        )

    lowered = message.lower()
    if 'excel file format cannot be determined' in lowered:
        return make_error_payload(
            'Planilha inválida',
            'Não foi possível identificar a estrutura da planilha enviada.',
            'Tente salvar novamente o arquivo como .xlsx ou envie em .csv.'
        )

    if 'unicode' in lowered or 'codec' in lowered or 'encoding' in lowered:
        return make_error_payload(
            'Erro de leitura do arquivo',
            'Não conseguimos interpretar a codificação do arquivo enviado.',
            'Tente abrir e salvar novamente o arquivo como CSV UTF-8 ou planilha .xlsx.'
        )

    if context == 'upload':
        return make_error_payload(
            'Falha ao ler o arquivo',
            'Não foi possível abrir esse arquivo para montar a prévia.',
            'Confira se o arquivo não está corrompido e tente novamente.',
            details=message
        )

    if context == 'process':
        return make_error_payload(
            'Falha na higienização',
            'O arquivo foi recebido, mas houve um problema durante o processamento dos dados.',
            'Revise o mapeamento das colunas e tente novamente.',
            details=message
        )

    if context == 'download':
        return make_error_payload(
            'Falha ao gerar o download',
            'Não foi possível montar o arquivo final higienizado.',
            'Tente processar novamente a lista antes de baixar.',
            details=message
        )

    if context == 'preview':
        return make_error_payload(
            'Falha na visualização',
            'Não foi possível atualizar a prévia com os filtros atuais.',
            'Tente alterar os filtros ou recarregar o arquivo.',
            details=message
        )

    return make_error_payload(
        'Erro inesperado',
        'Ocorreu um erro que não esperávamos neste fluxo.',
        'Tente novamente. Se persistir, revise o arquivo e os dados de entrada.',
        details=message
    )


def jsonify_error(exc: Exception, status_code: int, context: str):
    return jsonify({'error': classify_exception(exc, context)}), status_code


def _cache_put(cache: OrderedDict, key: str, value: dict, max_size: int) -> None:
    if key in cache:
        cache.pop(key)
    cache[key] = value
    while len(cache) > max_size:
        cache.popitem(last=False)


def store_uploaded_dataframe(filename: str, df: pd.DataFrame) -> str:
    token = uuid.uuid4().hex
    _cache_put(UPLOAD_CACHE, token, {'filename': filename, 'df': df}, UPLOAD_CACHE_MAX)
    return token


def get_uploaded_dataframe(token: str) -> dict | None:
    cached = UPLOAD_CACHE.get(token)
    if not cached:
        return None
    UPLOAD_CACHE.move_to_end(token)
    return cached


def build_result_cache_key(upload_token: str | None, filename: str, config: dict) -> str | None:
    if not upload_token:
        return None
    config_key = json.dumps(config, sort_keys=True, ensure_ascii=False)
    return f'{upload_token}:{filename}:{config_key}'


def store_processed_result(cache_key: str | None, result: dict) -> None:
    if not cache_key:
        return
    _cache_put(RESULT_CACHE, cache_key, result, RESULT_CACHE_MAX)


def get_processed_result(cache_key: str | None) -> dict | None:
    if not cache_key:
        return None
    cached = RESULT_CACHE.get(cache_key)
    if not cached:
        return None
    RESULT_CACHE.move_to_end(cache_key)
    return cached


def load_dataframe_from_request(req) -> tuple[pd.DataFrame, str, str | None]:
    upload_token = req.form.get('upload_token')
    if upload_token:
        cached = get_uploaded_dataframe(upload_token)
        if cached:
            return cached['df'], cached['filename'], upload_token

    file = req.files.get('file')
    if not file:
        raise ValueError('Nenhum arquivo enviado')

    df = read_file(file).fillna('')
    if len(df.columns) == 0:
        raise ValueError('Arquivo sem colunas legíveis')
    return df, file.filename, None


def apply_row_range(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    if not config.get('use_row_range'):
        return df

    total_rows = len(df)
    start_line = max(1, int(config.get('row_start', 1) or 1))
    end_line = max(start_line, int(config.get('row_end', total_rows) or total_rows))
    start_idx = min(start_line - 1, total_rows)
    end_idx = min(end_line, total_rows)
    return df.iloc[start_idx:end_idx].copy()


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.fillna('').astype(str).replace({'nan': '', 'None': '', '<NA>': ''})
    cleaned = cleaned.apply(lambda col: col.str.strip())
    return cleaned.loc[cleaned.ne('').any(axis=1)].reset_index(drop=True)


def get_or_process_result(df: pd.DataFrame, filename: str, upload_token: str | None, config: dict) -> dict:
    result_cache_key = build_result_cache_key(upload_token, filename, config)
    result = get_processed_result(result_cache_key)
    if result is None:
        scoped_df = apply_row_range(df, config)
        result = process_dataframe(scoped_df, config)
        store_processed_result(result_cache_key, result)
    return result


def parse_datetime_series(series: pd.Series) -> pd.Series:
    cleaned = series.fillna('').astype(str).str.strip()
    parsed = pd.Series(pd.NaT, index=cleaned.index, dtype='datetime64[ns]')
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        pending = parsed.isna() & cleaned.ne('')
        if not pending.any():
            break
        parsed.loc[pending] = pd.to_datetime(cleaned.loc[pending], format=fmt, errors='coerce')
    return parsed


def infer_date_columns(df: pd.DataFrame, sample_size: int = 200) -> list[str]:
    date_cols = []
    for col in df.columns:
        series = df[col].fillna('').astype(str).str.strip()
        sample = series[series.ne('')].head(sample_size)
        if sample.empty:
            continue
        parsed = parse_datetime_series(sample)
        hits = int(parsed.notna().sum())
        if hits >= 2 and hits / len(sample) >= 0.35:
            date_cols.append(col)
    return date_cols


def build_preview_meta(df: pd.DataFrame, config: dict, duplicate_review: dict | None = None) -> dict:
    name_col = config.get('name_col') if config.get('name_col') in df.columns else None
    phone_col = config.get('phone_col') if config.get('phone_col') in df.columns else None
    date_cols = infer_date_columns(df)
    duplicate_review = duplicate_review or {}
    has_duplicate_filters = bool(duplicate_review.get('available'))
    has_duplicate_review = bool(duplicate_review.get('enabled'))
    duplicate_removed = bool(duplicate_review.get('removed'))
    duplicate_count = int(duplicate_review.get('count', 0) or 0)

    phone_filter_options = [{'value': 'all', 'label': 'Todos os registros'}]
    if phone_col:
        phone_filter_options.extend([
            {'value': 'phone_present', 'label': 'Com telefone'},
            {'value': 'phone_missing', 'label': 'Sem telefone'},
            {'value': 'phone_valid', 'label': 'Telefones válidos'},
            {'value': 'phone_invalid', 'label': 'Telefones inválidos'},
        ])
    if has_duplicate_filters:
        phone_filter_options.extend([
            {'value': 'dup_all', 'label': f'Grupo de duplicadas ({duplicate_count})'},
            {'value': 'dup_kept', 'label': 'Original do grupo'},
            {'value': 'dup_removed', 'label': 'Duplicadas removidas' if duplicate_removed else 'Duplicadas do grupo'},
        ])

    sort_options = [{'value': '', 'label': 'Sem ordenação'}]
    if name_col:
        sort_options.extend([
            {'value': f'name::{name_col}::asc', 'label': f'{name_col} A → Z'},
            {'value': f'name::{name_col}::desc', 'label': f'{name_col} Z → A'},
        ])
    if phone_col:
        sort_options.extend([
            {'value': f'ddd::{phone_col}::asc', 'label': 'DDD crescente'},
            {'value': f'ddd::{phone_col}::desc', 'label': 'DDD decrescente'},
        ])
    for col in date_cols:
        sort_options.extend([
            {'value': f'date::{col}::desc', 'label': f'{col} mais recente'},
            {'value': f'date::{col}::asc', 'label': f'{col} mais antiga'},
        ])

    return {
        'phone_col': phone_col,
        'name_col': name_col,
        'date_columns': date_cols,
        'phone_filter_options': phone_filter_options,
        'sort_options': sort_options,
        'duplicate_review': {
            'available': has_duplicate_filters,
            'enabled': has_duplicate_review,
            'count': duplicate_count,
        },
    }


def apply_record_filter(frame: pd.DataFrame, phone_col: str | None, phone_filter: str) -> pd.DataFrame:
    if phone_filter.startswith('dup_'):
        if 'STATUS_DUPLICATA' not in frame.columns:
            return frame.iloc[0:0]
        if phone_filter == 'dup_all':
            return frame
        if phone_filter == 'dup_kept':
            return frame.loc[frame['STATUS_DUPLICATA'].eq('Original mantida')]
        if phone_filter == 'dup_removed':
            return frame.loc[frame['STATUS_DUPLICATA'].isin(['Duplicada removida', 'Duplicada mantida'])]
        return frame

    if not phone_col or phone_col not in frame.columns or phone_filter == 'all':
        return frame

    series = frame[phone_col].fillna('').astype(str).str.strip()
    normalized = series.map(lambda val: looks_like_phone(val, reject_documents=False))
    has_value = series.ne('')
    is_valid = normalized.notna()

    if phone_filter == 'phone_present':
        return frame.loc[has_value]
    if phone_filter == 'phone_missing':
        return frame.loc[~has_value]
    if phone_filter == 'phone_valid':
        return frame.loc[is_valid]
    if phone_filter == 'phone_invalid':
        return frame.loc[has_value & ~is_valid]
    return frame


def apply_date_filter(frame: pd.DataFrame, date_col: str | None, date_from: str, date_to: str) -> pd.DataFrame:
    if not date_col or date_col not in frame.columns or (not date_from and not date_to):
        return frame

    parsed = parse_datetime_series(frame[date_col])
    mask = parsed.notna()
    if date_from:
        mask &= parsed >= pd.to_datetime(date_from)
    if date_to:
        mask &= parsed <= pd.to_datetime(date_to) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
    return frame.loc[mask]


def apply_preview_sort(frame: pd.DataFrame, sort_mode: str) -> pd.DataFrame:
    if not sort_mode:
        return frame

    parts = sort_mode.split('::', 2)
    if len(parts) != 3:
        return frame

    mode, col, direction = parts
    ascending = direction != 'desc'
    if col not in frame.columns:
        return frame

    if mode == 'name':
        sort_key = frame[col].fillna('').astype(str).str.strip().str.casefold()
    elif mode == 'ddd':
        normalized = frame[col].fillna('').astype(str).map(
            lambda val: looks_like_phone(val, reject_documents=False)
        )
        sort_key = pd.to_numeric(normalized.str[2:4], errors='coerce')
    elif mode == 'date':
        sort_key = parse_datetime_series(frame[col])
    else:
        return frame

    return frame.assign(_sort_key=sort_key).sort_values(
        by=['_sort_key', col], ascending=ascending, kind='mergesort', na_position='last'
    ).drop(columns=['_sort_key'])


def build_preview_payload(df: pd.DataFrame, cols: list[str], config: dict, search: str = '', limit: int = 10,
                          phone_filter: str = 'all', sort_mode: str = '',
                          date_col: str | None = None, date_from: str = '', date_to: str = '',
                          duplicate_review: dict | None = None) -> dict:
    frame = df.copy()
    frame = frame.astype(str).replace({'nan': '', 'None': '', '<NA>': ''})
    total_rows = len(frame)
    meta = build_preview_meta(frame, config, duplicate_review)

    search = str(search or '').strip()
    if search:
        escaped = re.escape(search)
        mask = frame.apply(lambda col: col.str.contains(escaped, case=False, na=False))
        frame = frame.loc[mask.any(axis=1)]

    frame = apply_record_filter(frame, meta.get('phone_col'), phone_filter)
    frame = apply_date_filter(frame, date_col, date_from, date_to)
    filtered_total = len(frame)
    frame = apply_preview_sort(frame, sort_mode)

    limit = max(1, min(int(limit or 10), 200))
    frame = frame.head(limit)

    return {
        'columns': cols,
        'rows': frame[cols].to_dict(orient='records'),
        'shown_count': len(frame),
        'filtered_total': filtered_total,
        'total_rows': total_rows,
        'meta': meta,
    }


# ─── File reading ──────────────────────────────────────────────────────────────

def detect_csv_delimiter(sample_text: str) -> str:
    sample_text = str(sample_text or '').strip()
    if not sample_text:
        return ','

    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=',;|\t')
        if dialect.delimiter:
            return dialect.delimiter
    except csv.Error:
        pass

    candidates = {',': 0, ';': 0, '\t': 0, '|': 0}
    lines = [line for line in sample_text.splitlines()[:10] if line.strip()]
    for line in lines:
        for delimiter in candidates:
            candidates[delimiter] += line.count(delimiter)

    best_delimiter = max(candidates, key=candidates.get)
    return best_delimiter if candidates[best_delimiter] > 0 else ','


def read_excel_rows(buf: io.BytesIO) -> pd.DataFrame:
    import openpyxl

    buf.seek(0)
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    headers = None
    empty_streak = 0
    saw_data = False
    max_empty_streak = 2000

    for row in ws.iter_rows(values_only=True):
        values = [str(cell).strip() if cell is not None else '' for cell in row]
        if not any(values):
            if saw_data:
                empty_streak += 1
                if empty_streak >= max_empty_streak:
                    break
            continue

        empty_streak = 0
        saw_data = True

        if headers is None:
            headers = values
            continue

        if len(values) < len(headers):
            values += [''] * (len(headers) - len(values))
        elif len(values) > len(headers):
            values = values[:len(headers)]
        rows.append(values)

    if headers is None:
        return pd.DataFrame()
    return pd.DataFrame(rows, columns=headers).astype(str)


def read_file(file) -> pd.DataFrame:
    name = file.filename.lower()
    raw = file.read()
    buf = io.BytesIO(raw)
    if name.endswith('.csv'):
        for enc in ('utf-8', 'latin-1', 'cp1252'):
            try:
                sample = raw[:8192].decode(enc, errors='ignore')
                delimiter = detect_csv_delimiter(sample)
                buf.seek(0)
                df = pd.read_csv(buf, encoding=enc, dtype=str, sep=delimiter)
                return normalize_imported_dataframe(restore_headerless_first_row(df))
            except Exception:
                pass
    elif name.endswith(('.xlsx', '.xls')):
        return normalize_imported_dataframe(restore_headerless_first_row(read_excel_rows(buf)))
    elif name.endswith('.pdf'):
        import pdfplumber
        rows = []
        buf.seek(0)
        with pdfplumber.open(buf) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables():
                    for row in table:
                        rows.append([str(c) if c else '' for c in row])
        if not rows:
            return pd.DataFrame()
        pdf_df = pd.DataFrame(rows[1:], columns=rows[0]).astype(str)
        return normalize_imported_dataframe(restore_headerless_first_row(pdf_df))
    raise ValueError(f'Formato não suportado: {name}')


# ─── Detect misplaced phones ───────────────────────────────────────────────────

def detect_misplaced_phones(df: pd.DataFrame, phone_col: str, keep_cols: list) -> list:
    """
    For rows where phone_col is empty or invalid, scan other columns for phone-like values.
    Returns list of dicts: {row_index, name, found_col, found_value, normalized}
    """
    if phone_col not in df.columns:
        return []

    suggestions = []
    other_cols = [c for c in df.columns if c != phone_col and not is_document_column_name(c)]
    phone_series = df[phone_col].fillna('').astype(str).str.strip()
    pending_rows = phone_series.map(lambda val: looks_like_phone(val, reject_documents=False) is None)
    if not pending_rows.any():
        return []

    name_series = df.iloc[:, 0].fillna('').astype(str)
    found_frames = []
    for col in other_cols:
        candidate_values = df.loc[pending_rows, col].fillna('').astype(str).str.strip()
        if candidate_values.empty:
            continue

        normalized = candidate_values.map(looks_like_phone)
        matched = normalized.notna()
        if not matched.any():
            continue

        matched_index = candidate_values.index[matched]
        found_frames.append(pd.DataFrame({
            'row_index': matched_index.astype(int),
            'name': name_series.loc[matched_index].where(
                name_series.loc[matched_index].str.strip().ne(''),
                [f'Linha {idx+2}' for idx in matched_index]
            ),
            'found_col': col,
            'found_value': candidate_values.loc[matched_index].values,
            'normalized': normalized.loc[matched_index].values,
        }))
        pending_rows.loc[matched_index] = False
        if not pending_rows.any():
            break

    if found_frames:
        suggestions = (
            pd.concat(found_frames, ignore_index=True)
            .sort_values('row_index')
            .to_dict(orient='records')
        )
    return suggestions


# ─── Core processing (vectorized) ─────────────────────────────────────────────

def process_dataframe(df: pd.DataFrame, config: dict) -> dict:
    name_col      = config.get('name_col')
    phone_col     = config.get('phone_col')
    split_names   = config.get('split_names', True)
    dup_action    = config.get('dup_action', 'keep')
    keep_cols     = config.get('keep_cols', list(df.columns))
    # List of accepted misplaced phones from the UI
    phone_fixes   = {
        int(f['row_index']): {
            'normalized': f.get('normalized', ''),
            'found_col':  f.get('found_col'),
        }
        for f in config.get('phone_fixes', [])
        if 'row_index' in f
    }

    warnings = []
    original_total = len(df)
    metrics = {
        'original_rows': original_total,
        'empty_rows_removed': 0,
        'rows_with_multiple_phones': 0,
        'extra_rows_from_phone_split': 0,
        'invalid_phone_entries': 0,
        'rows_removed_without_phone': 0,
        'duplicates_found': 0,
        'duplicates_removed': 0,
        'final_rows': 0,
    }

    df = clean_dataframe(df)
    metrics['empty_rows_removed'] = max(original_total - len(df), 0)

    # Apply accepted phone fixes before processing
    if phone_col and phone_fixes:
        for row_idx, fix in phone_fixes.items():
            if row_idx in df.index:
                norm_phone = normalize_phone(str(fix.get('normalized', '')))
                if not norm_phone:
                    continue
                df.at[row_idx, phone_col] = norm_phone
                found_col = fix.get('found_col')
                if found_col and found_col in df.columns and found_col != phone_col:
                    df.at[row_idx, found_col] = ''

    # ── Handle multiple phones per row (explode) ──
    if phone_col and phone_col in df.columns:
        rows_before_split = len(df)
        df['_phones'] = df[phone_col].apply(
            lambda v: extract_phones(v) if v.strip() else ['']
        )
        if name_col and name_col in df.columns:
            cleaned_names = df[name_col].fillna('').astype(str).apply(normalize_person_name)
            parsed_names = df[name_col].fillna('').astype(str).apply(parse_list_cell)
            df['_paired_names'] = [
                [normalize_person_name(name) for name in names]
                if len(phones) > 1 and len(names) == len(phones) else
                [original_name] * len(phones)
                for original_name, names, phones in zip(cleaned_names, parsed_names, df['_phones'])
            ]
        phone_counts = df['_phones'].apply(len)
        metrics['rows_with_multiple_phones'] = int((phone_counts > 1).sum())
        explode_cols = ['_phones']
        if '_paired_names' in df.columns:
            explode_cols.append('_paired_names')
        df = df.explode(explode_cols).reset_index(drop=True)
        metrics['extra_rows_from_phone_split'] = max(len(df) - rows_before_split, 0)
        df['_phones'] = df['_phones'].fillna('').astype(str).replace(
            {'nan': '', 'None': '', 'NaN': ''})
        if '_paired_names' in df.columns:
            df[name_col] = df['_paired_names'].fillna('').astype(str)

        def _norm(v):
            if not isinstance(v, str):
                return None
            return normalize_phone(v)

        normalized = df['_phones'].apply(_norm)

        bad_mask = normalized.isna() & df['_phones'].str.strip().ne('')
        metrics['invalid_phone_entries'] = int(bad_mask.sum())
        for idx in df[bad_mask].index:
            warnings.append(f'Linha {idx+2}: telefone inválido "{df["_phones"][idx]}"')

        # Keep only normalized phones in the final file.
        df[phone_col] = normalized.fillna('')
        drop_cols = ['_phones']
        if '_paired_names' in df.columns:
            drop_cols.append('_paired_names')
        df = df.drop(columns=drop_cols)

    dedupe_basis = None

    # ── Select columns ──
    cols_to_keep = [c for c in keep_cols if c in df.columns]
    result = df[cols_to_keep].copy()

    # ── Name handling ──
    if name_col and name_col in result.columns:
        name_series = result[name_col].astype(str).apply(normalize_person_name)
        result[name_col] = name_series
        if split_names:
            split = name_series.str.split(n=1, expand=True)
            first_name = split[0].fillna('').str.capitalize() if 0 in split.columns else pd.Series('', index=result.index)
            last_name = split[1].fillna('').str.title() if 1 in split.columns else pd.Series('', index=result.index)
            name_pos = result.columns.get_loc(name_col)
            if 'first_name' in result.columns:
                result['first_name'] = first_name
            else:
                result.insert(name_pos + 1, 'first_name', first_name)
            if 'last_name' in result.columns:
                result['last_name'] = last_name
            else:
                name_pos = result.columns.get_loc(name_col)
                first_name_pos = result.columns.get_loc('first_name') if 'first_name' in result.columns else name_pos
                insert_pos = max(name_pos, first_name_pos) + 1
                result.insert(insert_pos, 'last_name', last_name)

    # ── Reorder columns ──
    ordered = []
    if name_col and name_col in result.columns:
        ordered.append(name_col)
    if split_names and name_col:
        if 'first_name' in result.columns: ordered.append('first_name')
        if 'last_name'  in result.columns: ordered.append('last_name')
    if phone_col and phone_col in result.columns:
        ordered.append(phone_col)
    for c in result.columns:
        if c not in ordered:
            ordered.append(c)
    result = result[[c for c in ordered if c in result.columns]]

    # ── Remove rows without phone ──
    if phone_col and phone_col in result.columns:
        rows_before_phone_filter = len(result)
        result = result[result[phone_col].astype(str).str.strip().ne('')]
        dedupe_basis = result[phone_col].fillna('').astype(str).str.strip()
        metrics['rows_removed_without_phone'] = max(rows_before_phone_filter - len(result), 0)
    else:
        dedupe_basis = pd.Series('', index=result.index, dtype=str)

    # ── Duplicates ──
    duplicate_review_df = pd.DataFrame()
    duplicate_group_mask = dedupe_basis.duplicated(keep=False)
    duplicate_row_mask = dedupe_basis.duplicated(keep='first')
    duplicate_group = result.loc[duplicate_group_mask].copy()
    if not duplicate_group.empty:
        duplicate_group.insert(
            0,
            'STATUS_DUPLICATA',
            np.where(
                duplicate_row_mask.loc[duplicate_group.index],
                'Duplicada removida' if dup_action == 'remove' else 'Duplicada mantida',
                'Original mantida'
            )
        )
        duplicate_review_df = duplicate_group

    dup_count = int(duplicate_row_mask.sum())
    metrics['duplicates_found'] = dup_count
    if dup_action == 'remove':
        before_drop_duplicates = len(result)
        keep_mask = ~duplicate_row_mask
        result = result.loc[keep_mask]
        dedupe_basis = dedupe_basis.loc[keep_mask]
        metrics['duplicates_removed'] = max(before_drop_duplicates - len(result), 0)
    metrics['final_rows'] = len(result)

    return {
        'df':        result,
        'dup_count': dup_count,
        'warnings':  warnings,
        'metrics':   metrics,
        'duplicate_review_df': duplicate_review_df,
    }


# ─── FakeFile helper ───────────────────────────────────────────────────────────

class FakeFile:
    def __init__(self, buf, name):
        self.filename = name
        self._buf = buf
    def read(self):
        self._buf.seek(0)
        return self._buf.read()
    def seek(self, *a):
        return self._buf.seek(*a)
    def tell(self):
        return self._buf.tell()


# ─── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file:
        return jsonify_error(ValueError('Nenhum arquivo enviado'), 400, 'upload')
    try:
        df = read_file(file)
        df = df.fillna('')
        if len(df.columns) == 0:
            raise ValueError('Arquivo sem colunas legíveis')
        upload_token = store_uploaded_dataframe(file.filename, df)
        inferred = infer_columns(df)
        return jsonify({
            'columns':           list(df.columns),
            'preview':           df.head(10).to_dict(orient='records'),
            'total_rows':        len(df),
            'suggested_columns': inferred,
            'upload_token':      upload_token,
        })
    except Exception as e:
        status_code = 400 if isinstance(e, (ValueError, pd.errors.EmptyDataError)) else 500
        return jsonify_error(e, status_code, 'upload')


@app.route('/process_stream', methods=['POST'])
def process_stream():
    config   = json.loads(request.form.get('config', '{}'))
    try:
        df, filename, upload_token = load_dataframe_from_request(request)
    except Exception as e:
        def generate_error():
            payload = {'error': classify_exception(e, 'process')}
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"
        return Response(stream_with_context(generate_error()), mimetype='text/event-stream',
                        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})

    def generate():
        def emit(pct, msg, data=None):
            payload = {'pct': pct, 'msg': msg}
            if data: payload['data'] = data
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

        try:
            yield from emit(5, 'Lendo arquivo…')
            scoped_df = apply_row_range(df, config)
            total = len(scoped_df)

            yield from emit(20, f'{total} linhas encontradas. Processando…')
            result = get_or_process_result(df, filename, upload_token, config)
            result_df = result['df']

            # Detect misplaced phones (only if no fixes already provided)
            phone_col = config.get('phone_col')
            misplaced = []
            if phone_col and not config.get('phone_fixes'):
                df_clean = clean_dataframe(scoped_df)
                misplaced = detect_misplaced_phones(df_clean, phone_col, config.get('keep_cols', []))

            yield from emit(85, 'Finalizando…')

            payload = {
                'columns_before': list(df.columns),
                'columns_after':  list(result_df.columns),
                'preview_before': scoped_df.head(10).to_dict(orient='records'),
                'preview_after':  result_df.head(10).to_dict(orient='records'),
                'total_before':   total,
                'total_after':    len(result_df),
                'dup_count':      result['dup_count'],
                'warnings':       result['warnings'],
                'metrics':        result.get('metrics', {}),
                'misplaced':      misplaced,
            }
            yield from emit(100, 'Concluído!', payload)

        except Exception as e:
            import traceback; traceback.print_exc()
            yield f"data: {json.dumps({'error': classify_exception(e, 'process')}, ensure_ascii=False)}\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/download', methods=['POST'])
def download():
    config = json.loads(request.form.get('config', '{}'))
    fmt    = request.form.get('format', 'xlsx')
    total_parts = max(1, int(request.form.get('total_parts', '1') or '1'))
    part_index = max(1, int(request.form.get('part_index', '1') or '1'))
    try:
        df, filename, upload_token = load_dataframe_from_request(request)
    except ValueError as e:
        return jsonify_error(e, 400, 'download')
    try:
        result = get_or_process_result(df, filename, upload_token, config)
        result_df = result['df']
        parts = split_dataframe_evenly(result_df, total_parts)
        if part_index > len(parts):
            raise ValueError('Parte de download inválida')
        result_df = parts[part_index - 1]
        download_name = (
            build_split_download_name(filename, fmt, part_index, total_parts)
            if total_parts > 1 else
            build_download_name(filename, fmt)
        )
        buf = io.BytesIO()
        if fmt == 'csv':
            result_df.to_csv(buf, index=False, encoding='utf-8-sig', sep=';')
            buf.seek(0)
            return send_file(buf, mimetype='text/csv',
                             as_attachment=True, download_name=download_name)
        else:
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Higienizado')
            buf.seek(0)
            return send_file(buf,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=download_name)
    except Exception as e:
        status_code = 400 if isinstance(e, (ValueError, pd.errors.EmptyDataError)) else 500
        return jsonify_error(e, status_code, 'download')


@app.route('/preview', methods=['POST'])
def preview():
    config = json.loads(request.form.get('config', '{}'))
    stage = request.form.get('stage', 'before')
    search = request.form.get('search', '')
    phone_filter = request.form.get('phone_filter', 'all')
    sort_mode = request.form.get('sort_mode', '')
    date_col = request.form.get('date_col') or None
    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    limit = request.form.get('limit', '10')

    try:
        df, filename, upload_token = load_dataframe_from_request(request)
    except ValueError as e:
        return jsonify_error(e, 400, 'preview')
    try:
        scoped_df = apply_row_range(df, config)
        if stage == 'after':
            result = get_or_process_result(df, filename, upload_token, config)
            duplicate_review_df = result.get('duplicate_review_df')
            duplicate_review = {
                'available': True,
                'enabled': bool(duplicate_review_df is not None and not duplicate_review_df.empty),
                'removed': bool(result.get('metrics', {}).get('duplicates_removed', 0)),
                'count': int(result.get('metrics', {}).get('duplicates_found', 0)),
            }
            if phone_filter.startswith('dup_') and duplicate_review['enabled']:
                preview_df = duplicate_review_df
            else:
                preview_df = result['df']
        else:
            preview_df = scoped_df.fillna('').astype(str)
            duplicate_review = {'available': False, 'enabled': False, 'removed': False, 'count': 0}

        cols = list(preview_df.columns)
        payload = build_preview_payload(
            preview_df,
            cols=cols,
            config=config,
            search=search,
            limit=limit,
            phone_filter=phone_filter,
            sort_mode=sort_mode,
            date_col=date_col,
            date_from=date_from,
            date_to=date_to,
            duplicate_review=duplicate_review,
        )
        return jsonify(payload)
    except Exception as e:
        status_code = 400 if isinstance(e, (ValueError, pd.errors.EmptyDataError)) else 500
        return jsonify_error(e, status_code, 'preview')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5050)
